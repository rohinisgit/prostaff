import calendar
import io
from datetime import date
from decimal import Decimal
import calendar
import io
from datetime import date, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile

from attendance.models import AttendanceRecord, MonthlyAttendanceSheet
def normalize_year_month(year, month):
    """Wraps month overflow/underflow into the correct year, e.g. month=13 -> next Jan."""
    year, month = int(year), int(month)
    while month > 12:
        month -= 12
        year += 1
    while month < 1:
        month += 12
        year -= 1
    return year, month


def get_cl_quota(user):
    """Fixed monthly CL quota shown on the sheet (separate from LeaveBalance).
    2 for Manager, 1 for everyone else (Employee/HR/Admin)."""
    if user.role == 'MANAGER':
        return 2
    return 1


def build_monthly_summary(user, year, month):
    """Builds the full monthly attendance summary dict for one employee."""
    year, month = normalize_year_month(year, month)
    days_in_month = calendar.monthrange(year, month)[1]
    first_day = date(year, month, 1)
    last_day = date(year, month, days_in_month)

    records = AttendanceRecord.objects.filter(user=user, date__gte=first_day, date__lte=last_day)
    records_by_date = {r.date: r for r in records}

    daily_rows = []
    sundays = 0
    on_duty_days = 0          # any day attended, including Sunday
    on_duty_non_sunday = 0    # attended days excluding Sunday (counts against WKD)
    sunday_hours = Decimal('0.00')
    night_duty_count = 0
    total_hours_worked = Decimal('0.00')

    for day_num in range(1, days_in_month + 1):
        current_date = date(year, month, day_num)
        is_sunday = current_date.weekday() == 6  # Monday=0 ... Sunday=6
        if is_sunday:
            sundays += 1

        record = records_by_date.get(current_date)
        hours = record.total_hours if record and record.total_hours else Decimal('0.00')
        attended = bool(record and record.in_time)

        if attended:
            on_duty_days += 1
            total_hours_worked += hours
            if not is_sunday:
                on_duty_non_sunday += 1

        if is_sunday and record:
            sunday_hours += hours

        if record and record.is_night_duty:
            night_duty_count += 1

        daily_rows.append({
            'date': current_date,
            'day_label': current_date.strftime('%d %b'),
            'is_sunday': is_sunday,
            'hours': hours,
            'in_time': record.in_time if record else None,
            'out_time': record.out_time if record else None,
            'attended': attended,
        })

    wkd = days_in_month - sundays  # working days excluding Sunday
    actual_leave = max(0, wkd - on_duty_non_sunday)  # WKD days not attended

    cl_quota = get_cl_quota(user)
    lop = max(0, actual_leave - cl_quota)

    return {
        'user': user,
        'year': year,
        'month': month,
        'month_name': first_day.strftime('%B %Y'),
        'days_in_month': days_in_month,
        'daily_rows': daily_rows,
        'total_days_worked': on_duty_days,
        'on_duty': on_duty_days,
        'cl_quota': cl_quota,
        'sunday_hours': sunday_hours,
        'total_hours_worked': total_hours_worked,
        'wkd': wkd,
        'actual_leave': actual_leave,
        'lop': lop,
        'total_ph_sunday': sundays,
        'night_duty': night_duty_count,
    }


def generate_monthly_excel(summary):
    """Builds an .xlsx workbook (in-memory) matching the monthly register
    format: one row per employee, one column per calendar day, plus the
    summary columns at the end."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{summary['month']:02d}-{summary['year']}"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0057FF", end_color="0057FF", fill_type="solid")
    sunday_fill = PatternFill(start_color="FFF1E8", end_color="FFF1E8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    user = summary['user']

    ws['A1'] = 'Employee Name'
    ws['B1'] = user.get_full_name() or user.username
    ws['A2'] = 'Employee ID'
    ws['B2'] = user.employee_id or '-'
    ws['A3'] = 'Month'
    ws['B3'] = summary['month_name']
    for cell in ('A1', 'A2', 'A3'):
        ws[cell].font = bold

    header_row = 5
    col = 1
    ws.cell(row=header_row, column=col, value='Employee').font = bold
    col += 1
    ws.cell(row=header_row, column=col, value='Emp ID').font = bold
    col += 1

    day_start_col = col
    for day in summary['daily_rows']:
        c = ws.cell(row=header_row, column=col, value=day['day_label'])
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        col += 1
    day_end_col = col

    summary_headers = [
        'Total Days Worked', 'CL', 'On Duty', 'Sunday Hours', 'Total Hours',
        'WKD (Excl. Sunday)', 'Actual Leave', 'LOP', 'Total (PH/Sunday)',
        'Total Days in Month', 'Night Duty',
    ]
    summary_start_col = col
    for h in summary_headers:
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        col += 1
    summary_end_col = col

    data_row = header_row + 1
    col = 1
    c = ws.cell(row=data_row, column=col, value=user.get_full_name() or user.username)
    c.border = border
    col += 1
    c = ws.cell(row=data_row, column=col, value=user.employee_id or '-')
    c.border = border
    col += 1

    for day in summary['daily_rows']:
        val = float(day['hours']) if day['hours'] else 0
        c = ws.cell(row=data_row, column=col, value=val)
        c.alignment = center
        c.border = border
        if day['is_sunday']:
            c.fill = sunday_fill
        col += 1

    summary_values = [
        summary['total_days_worked'],
        summary['cl_quota'],
        summary['on_duty'],
        float(summary['sunday_hours']),
        float(summary['total_hours_worked']),
        summary['wkd'],
        summary['actual_leave'],
        summary['lop'],
        summary['total_ph_sunday'],
        summary['days_in_month'],
        summary['night_duty'],
    ]
    for v in summary_values:
        c = ws.cell(row=data_row, column=col, value=v)
        c.alignment = center
        c.border = border
        col += 1

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    for i in range(day_start_col, day_end_col):
        ws.column_dimensions[get_column_letter(i)].width = 9
    for i in range(summary_start_col, summary_end_col):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws.freeze_panes = ws.cell(row=header_row + 1, column=day_start_col)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def regenerate_and_save_monthly_sheet(user, year, month):
    """Recomputes the summary for this user/month and re-saves the Excel
    snapshot. Called automatically on every punch."""
    year, month = normalize_year_month(year, month)
    summary = build_monthly_summary(user, year, month)
    buffer = generate_monthly_excel(summary)

    sheet, _ = MonthlyAttendanceSheet.objects.get_or_create(user=user, year=year, month=month)
    filename = f"attendance_{user.username}_{year}_{month:02d}.xlsx"
    sheet.excel_file.save(filename, ContentFile(buffer.read()), save=True)
    return sheet

def build_team_monthly_summary(users, year, month):
    """Builds the monthly summary for every user in `users`, for the same
    month. Returns a list of summary dicts (same shape as
    build_monthly_summary), one per employee."""
    year, month = normalize_year_month(year, month)
    summaries = []
    for u in users:
        summaries.append(build_monthly_summary(u, year, month))
    return summaries


def generate_team_monthly_excel(summaries, year, month):
    """Builds one workbook with ONE ROW PER EMPLOYEE for the given month —
    a combined muster/register sheet, not per-employee tabs."""
    wb = Workbook()
    ws = wb.active

    if not summaries:
        ws.title = f"{month:02d}-{year}"
        ws['A1'] = 'No active employees found.'
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    ws.title = f"{month:02d}-{year}"[:31]

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0057FF", end_color="0057FF", fill_type="solid")
    sunday_fill = PatternFill(start_color="FFF1E8", end_color="FFF1E8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    days_in_month = summaries[0]['days_in_month']
    month_name = summaries[0]['month_name']

    ws['A1'] = 'Monthly Attendance Register'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = month_name
    ws['A2'].font = bold

    header_row = 4
    col = 1
    fixed_headers = ['Employee', 'Emp ID', 'Role', 'Department']
    for h in fixed_headers:
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        col += 1

    day_start_col = col
    day_labels = [d['day_label'] for d in summaries[0]['daily_rows']]

def get_week_options(year, month):
    """List of (week_number, label) for the week dropdown, splitting the
    month into consecutive 7-day chunks starting on the 1st."""
    year, month = normalize_year_month(year, month)
    days_in_month = calendar.monthrange(year, month)[1]
    options = []
    week_number = 1
    start_day = 1
    while start_day <= days_in_month:
        end_day = min(start_day + 6, days_in_month)
        label = f"Week {week_number} ({start_day}-{end_day} {date(year, month, 1).strftime('%b')})"
        options.append((week_number, label))
        start_day += 7
        week_number += 1
    return options


def build_weekly_summary(user, year, month, week_number):
    year, month = normalize_year_month(year, month)
    days_in_month = calendar.monthrange(year, month)[1]

    start_day = (week_number - 1) * 7 + 1
    if start_day > days_in_month:
        week_number = 1
        start_day = 1
    end_day = min(start_day + 6, days_in_month)

    first_day = date(year, month, start_day)
    last_day = date(year, month, end_day)

    records = AttendanceRecord.objects.filter(user=user, date__gte=first_day, date__lte=last_day)
    records_by_date = {r.date: r for r in records}

    daily_rows = []
    total_hours = Decimal('0.00')
    days_worked = 0
    sundays = 0
    night_duty_count = 0

    current = first_day
    while current <= last_day:
        is_sunday = current.weekday() == 6
        if is_sunday:
            sundays += 1
        record = records_by_date.get(current)
        hours = record.total_hours if record and record.total_hours else Decimal('0.00')
        attended = bool(record and record.in_time)
        if attended:
            days_worked += 1
            total_hours += hours
        if record and record.is_night_duty:
            night_duty_count += 1
        daily_rows.append({
            'date': current,
            'day_label': current.strftime('%d %b (%a)'),
            'is_sunday': is_sunday,
            'hours': hours,
            'in_time': record.in_time if record else None,
            'out_time': record.out_time if record else None,
            'attended': attended,
        })
        current += timedelta(days=1)

    return {
        'user': user, 'year': year, 'month': month, 'week_number': week_number,
        'start_date': first_day, 'end_date': last_day,
        'label': f"Week {week_number}: {first_day.strftime('%d %b')} - {last_day.strftime('%d %b %Y')}",
        'daily_rows': daily_rows,
        'days_worked': days_worked,
        'total_hours_worked': total_hours,
        'sundays': sundays,
        'night_duty': night_duty_count,
        'total_days': (last_day - first_day).days + 1,
    }


def build_yearly_summary(user, year):
    months_data = []
    total_hours = Decimal('0.00')
    total_days_worked = 0
    total_actual_leave = 0
    total_lop = 0
    total_night_duty = 0
    total_cl_quota = 0

    for m in range(1, 13):
        s = build_monthly_summary(user, year, m)
        months_data.append(s)
        total_hours += s['total_hours_worked']
        total_days_worked += s['total_days_worked']
        total_actual_leave += s['actual_leave']
        total_lop += s['lop']
        total_night_duty += s['night_duty']
        total_cl_quota += s['cl_quota']

    return {
        'user': user, 'year': year,
        'months': months_data,
        'total_hours_worked': total_hours,
        'total_days_worked': total_days_worked,
        'total_actual_leave': total_actual_leave,
        'total_lop': total_lop,
        'total_night_duty': total_night_duty,
        'total_cl_quota': total_cl_quota,
    }


def generate_weekly_excel(summary):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {summary['week_number']}"
    user = summary['user']
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0057FF", end_color="0057FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws['A1'] = 'Employee'; ws['B1'] = user.get_full_name() or user.username
    ws['A2'] = 'Week'; ws['B2'] = summary['label']
    ws['A1'].font = bold
    ws['A2'].font = bold

    headers = ['Date', 'Day', 'In Time', 'Out Time', 'Hours']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = header_fill

    row = 5
    for d in summary['daily_rows']:
        ws.cell(row=row, column=1, value=d['date'].strftime('%d-%m-%Y'))
        ws.cell(row=row, column=2, value=d['date'].strftime('%A'))
        ws.cell(row=row, column=3, value=d['in_time'].strftime('%I:%M %p') if d['in_time'] else '-')
        ws.cell(row=row, column=4, value=d['out_time'].strftime('%I:%M %p') if d['out_time'] else '-')
        ws.cell(row=row, column=5, value=float(d['hours']))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Total Hours').font = bold
    ws.cell(row=row, column=2, value=float(summary['total_hours_worked']))
    row += 1
    ws.cell(row=row, column=1, value='Days Worked').font = bold
    ws.cell(row=row, column=2, value=summary['days_worked'])

    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_yearly_excel(summary):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{summary['year']}"
    user = summary['user']
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="0057FF", end_color="0057FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws['A1'] = 'Employee'; ws['B1'] = user.get_full_name() or user.username
    ws['A2'] = 'Year'; ws['B2'] = summary['year']
    ws['A1'].font = bold
    ws['A2'].font = bold

    headers = ['Month', 'Days Worked', 'CL Quota', 'Total Hours', 'Actual Leave', 'LOP', 'Night Duty']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = header_fill

    row = 5
    for s in summary['months']:
        ws.cell(row=row, column=1, value=s['month_name'])
        ws.cell(row=row, column=2, value=s['total_days_worked'])
        ws.cell(row=row, column=3, value=s['cl_quota'])
        ws.cell(row=row, column=4, value=float(s['total_hours_worked']))
        ws.cell(row=row, column=5, value=s['actual_leave'])
        ws.cell(row=row, column=6, value=s['lop'])
        ws.cell(row=row, column=7, value=s['night_duty'])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='TOTAL').font = bold
    ws.cell(row=row, column=2, value=summary['total_days_worked']).font = bold
    ws.cell(row=row, column=3, value=summary['total_cl_quota']).font = bold
    ws.cell(row=row, column=4, value=float(summary['total_hours_worked'])).font = bold
    ws.cell(row=row, column=5, value=summary['total_actual_leave']).font = bold
    ws.cell(row=row, column=6, value=summary['total_lop']).font = bold
    ws.cell(row=row, column=7, value=summary['total_night_duty']).font = bold

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer